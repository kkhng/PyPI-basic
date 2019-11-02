# [Basic Python] Project 1: Automation with Python
# Use openpyxl to process spreadsheet
# Task1: The correct price for wrong_price in spreadsheet is 90% of the wrong_price, automate it!
import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
rename_cell = sheet['d1']
rename_cell.value = 'corrected_price'
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = round(corrected_price, 2)

wb.save('transactions.xlsx')