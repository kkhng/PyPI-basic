# [Basic Python] Project 1: Automation with Python
# Use openpyxl to process spreadsheet
# Task3: Create a reusable function using the codes you just written in pipt2 python file
import openpyxl as xl
from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.reference import Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    rename_cell = sheet['d1']
    rename_cell.value = 'corrected_price'
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = round(corrected_price, 2)

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = LineChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
